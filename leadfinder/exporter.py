"""Exports the SQLite database to a readable, filterable Excel workbook."""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import utils
from .database import Database

logger = utils.get_logger()

HEADERS = [
    "Lead Score", "Website Score", "Business", "Town", "Industry", "Website",
    "Email", "Phone", "Address", "Company Type", "Company Number",
    "Website Status", "HTTPS", "Mobile", "Load Time (s)", "Broken Links",
    "Contact Form", "Main Lead Reason", "All Lead Reasons", "Source",
    "Source URL", "Last Checked", "Screenshot Path", "Suppressed",
]

MAX_COL_WIDTH = 60
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUPPRESSED_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")


@dataclass
class LeadRow:
    business_id: int
    lead_score: Optional[int]
    website_score: Optional[int]
    business: str
    town: str
    industry: str
    website: Optional[str]
    email: Optional[str]
    phone: Optional[str]
    address: str
    company_type: str
    company_number: str
    website_status: str
    https: str
    mobile: str
    load_time: Optional[float]
    broken_links: Optional[int]
    contact_form: str
    main_reason: str
    all_reasons: str
    source: str
    source_url: str
    last_checked: str
    screenshot_path: str
    suppressed: bool


GENERIC_PREFIXES = (
    "info@", "hello@", "enquiries@", "enquiry@", "sales@", "office@", "contact@",
)


def _pick_display_email(contacts: list) -> Optional[str]:
    emails = [c["value"] for c in contacts if c["contact_type"] == "EMAIL"]
    if not emails:
        return None
    for prefix in GENERIC_PREFIXES:
        for e in emails:
            if e.lower().startswith(prefix):
                return e
    domain_emails = [c["value"] for c in contacts
                      if c["contact_type"] == "EMAIL" and c["email_classification"] == "DOMAIN_EMAIL"]
    if domain_emails:
        return domain_emails[0]
    return emails[0]


def build_lead_rows(db: Database) -> list[LeadRow]:
    rows: list[LeadRow] = []
    for biz in db.iter_businesses():
        website = db.get_website(biz["id"])
        scan = db.get_latest_scan(website["id"]) if website else None
        contacts = db.get_contacts(biz["id"])
        ch_match = db.get_ch_match(biz["id"])
        source = db.get_primary_source(biz["id"])
        screenshot = db.get_latest_screenshot(biz["id"])

        email = _pick_display_email(contacts)
        phone = biz["phone"] or next(
            (c["value"] for c in contacts if c["contact_type"] == "PHONE"), None
        )

        domain = utils.registrable_domain(utils.extract_domain(website["url"])) \
            if website and website["url"] else None
        all_emails = [c["value"] for c in contacts if c["contact_type"] == "EMAIL"]
        suppressed = db.is_suppressed(biz["name"], domain, all_emails)

        reasons = json.loads(biz["lead_reasons"]) if biz["lead_reasons"] else []
        company_type = "Limited Company" if biz["is_limited_company"] else \
            ("Sole trader / unincorporated (assumed)" if biz["is_limited_company"] is not None else "Unknown")

        rows.append(LeadRow(
            business_id=biz["id"],
            lead_score=biz["lead_score"],
            website_score=biz["website_score"],
            business=biz["name"],
            town=biz["town"] or "",
            industry=biz["industry_searched"] or biz["category"] or "",
            website=website["url"] if website else None,
            email=email,
            phone=phone,
            address=biz["address"] or "",
            company_type=company_type,
            company_number=ch_match["company_number"] if ch_match else (biz["companies_house_number"] or ""),
            website_status=website["status"] if website else "NOT_CHECKED",
            https="Yes" if scan and scan["https"] else ("No" if scan else ""),
            mobile="OK" if scan and scan["mobile_friendly"] else ("Issues" if scan else ""),
            load_time=scan["load_time_seconds"] if scan else None,
            broken_links=scan["broken_internal_links"] if scan else None,
            contact_form="Yes" if scan and scan["has_contact_form"] else ("No" if scan else ""),
            main_reason=reasons[0] if reasons else "",
            all_reasons="; ".join(reasons),
            source=source["source_name"] if source else "",
            source_url=source["source_url"] if source else "",
            last_checked=biz["last_checked"] or "",
            screenshot_path=screenshot["desktop_path"] if screenshot else "",
            suppressed=suppressed,
        ))
    return rows


def _write_sheet(wb: Workbook, title: str, rows: list[LeadRow]) -> Worksheet:
    ws = wb.create_sheet(title)
    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([
            row.lead_score, row.website_score, row.business, row.town, row.industry,
            row.website or "", row.email or "", row.phone or "", row.address,
            row.company_type, row.company_number, row.website_status, row.https,
            row.mobile, row.load_time, row.broken_links, row.contact_form,
            row.main_reason, row.all_reasons, row.source, row.source_url,
            row.last_checked, row.screenshot_path, "Yes" if row.suppressed else "",
        ])
        r = ws.max_row
        if row.website:
            ws.cell(row=r, column=6).hyperlink = row.website
            ws.cell(row=r, column=6).style = "Hyperlink"
        if row.source_url:
            ws.cell(row=r, column=21).hyperlink = row.source_url
            ws.cell(row=r, column=21).style = "Hyperlink"
        if row.suppressed:
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=col_idx).fill = SUPPRESSED_FILL

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
        ws.conditional_formatting.add(
            f"A2:A{ws.max_row}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            ),
        )

    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            value = row_cells[0].value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, MAX_COL_WIDTH)

    return ws


def export_to_excel(db: Database, output_path: Path, best_lead_threshold: int = 60,
                     poor_website_threshold: int = 50) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    all_rows = build_lead_rows(db)
    actionable = [r for r in all_rows if not r.suppressed]

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    best_leads = sorted(
        [r for r in actionable if (r.lead_score or 0) >= best_lead_threshold],
        key=lambda r: (r.lead_score or 0), reverse=True,
    )
    no_website = sorted(
        [r for r in actionable if r.website_status == "NO_WEBSITE_FOUND"],
        key=lambda r: (r.lead_score or 0), reverse=True,
    )
    poor_websites = sorted(
        [r for r in actionable if r.website_score is not None and r.website_score < poor_website_threshold],
        key=lambda r: (r.website_score or 0),
    )
    all_businesses = sorted(all_rows, key=lambda r: (r.lead_score or 0), reverse=True)
    dnc_rows = db.list_do_not_contact()

    _write_sheet(wb, "Best Leads", best_leads)
    _write_sheet(wb, "No Website", no_website)
    _write_sheet(wb, "Poor Websites", poor_websites)
    _write_sheet(wb, "All Businesses", all_businesses)

    dnc_ws = wb.create_sheet("Do Not Contact")
    dnc_headers = ["Business Name", "Domain", "Email", "Reason", "Date Added"]
    dnc_ws.append(dnc_headers)
    for col_idx in range(1, len(dnc_headers) + 1):
        cell = dnc_ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    dnc_ws.freeze_panes = "A2"
    for row in dnc_rows:
        dnc_ws.append([row["business_name"], row["domain"], row["email"], row["reason"], row["date_added"]])
    if dnc_ws.max_row > 1:
        dnc_ws.auto_filter.ref = f"A1:{get_column_letter(len(dnc_headers))}{dnc_ws.max_row}"
    for col_idx, header in enumerate(dnc_headers, start=1):
        max_len = len(header)
        for row_cells in dnc_ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            value = row_cells[0].value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        dnc_ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, MAX_COL_WIDTH)

    wb.save(str(output_path))
    logger.info("Exported %d business(es) to %s", len(all_rows), output_path)
    return output_path
